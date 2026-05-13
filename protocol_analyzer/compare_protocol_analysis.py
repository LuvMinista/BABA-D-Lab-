"""
Protocol Security Analysis Comparison: LLM vs Manual OFMC Results
=================================================================
 
Compares LLM-generated security verdicts (summary CSV) against
manually-run OFMC tool results (sessions=1 and sessions=2 CSVs).
 
Comparisons per protocol and goal:
  - Attack verdict agreement (LLM vs OFMC)
  - Disagreement classification (False Positive / False Negative)
  - Goal description side-by-side
  - Execution time (LLM duration vs OFMC execution-time)
  - Narrative interpretation of every difference
 
Output: Excel report with multiple sheets for academic/reporting use.
"""
 
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 1. File paths
# ─────────────────────────────────────────────────────────────────────────────
SUMMARY_CSV  = "./results/summary_20260320_211652.csv"
MANUAL_S2    = "./results/results-single-anb-typed-sessions_1-depth_0-2022-PCM.csv"
MANUAL_S1    = "./results/results-single-anb-typed-sessions_1-depth_0-2022-PCM.csv"
OUTPUT_PATH  = "./results/protocol_comparison_report.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# 2.  Colour / style constants
# ─────────────────────────────────────────────────────────────────────────────
C_HEADER     = PatternFill('solid', fgColor='1F4E79')
C_ALT        = PatternFill('solid', fgColor='D6E4F0')
C_GREEN      = PatternFill('solid', fgColor='C6EFCE')
C_RED        = PatternFill('solid', fgColor='FFC7CE')
C_YELLOW     = PatternFill('solid', fgColor='FFEB9C')
C_ORANGE     = PatternFill('solid', fgColor='FFD580')
C_BLUE_LIGHT = PatternFill('solid', fgColor='BDD7EE')
C_PURPLE     = PatternFill('solid', fgColor='E2CFEA')
C_GREY       = PatternFill('solid', fgColor='F2F2F2')
 
F_HEADER = Font(color='FFFFFF', bold=True, size=11)
F_BOLD   = Font(bold=True)
F_NORMAL = Font(size=10)
F_ITALIC = Font(italic=True, size=10)
 
THIN = Border(
    left=Side(style='thin',   color='B0C4DE'),
    right=Side(style='thin',  color='B0C4DE'),
    top=Side(style='thin',    color='B0C4DE'),
    bottom=Side(style='thin', color='B0C4DE'),
)
THICK_BOTTOM = Border(bottom=Side(style='medium', color='1F4E79'))
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 3.  Helpers
# ─────────────────────────────────────────────────────────────────────────────
 
def normalise_llm_protocol(name):
    return re.sub(r'_AnB\.AnBx$', '', str(name), flags=re.IGNORECASE).strip()
 
 
def parse_manual_protocol(name):
    m = re.match(r'^(.+?)_(\d+)$', str(name))
    return (m.group(1), int(m.group(2)) - 1) if m else (name, None)
 
 
def classify_disagreement(llm_v, ofmc_v):
    if pd.isna(llm_v) or pd.isna(ofmc_v):
        return 'Missing Data'
    if llm_v == ofmc_v == 'ATTACK':
        return 'True Positive – Both detected attack'
    if llm_v == ofmc_v == 'NO_ATTACK':
        return 'True Negative – Both agreed: no attack'
    if llm_v == 'ATTACK' and ofmc_v == 'NO_ATTACK':
        return 'False Positive – LLM raised attack; OFMC found none'
    if llm_v == 'NO_ATTACK' and ofmc_v == 'ATTACK':
        return 'False Negative – LLM missed attack found by OFMC'
    return 'Unknown'
 
 
def verdict_interpretation(llm_v, ofmc_v, llm_goal, ofmc_goal, reasoning):
    g = str(llm_goal).strip() if pd.notna(llm_goal) else '(unknown goal)'
    r = str(reasoning).strip() if pd.notna(reasoning) else ''
    if llm_v == ofmc_v == 'NO_ATTACK':
        return (f'Both the LLM and OFMC agree that the goal "{g}" is satisfied '
                f'with no attack possible. The protocol correctly protects this property.')
    if llm_v == ofmc_v == 'ATTACK':
        return (f'Both the LLM and OFMC independently identified an attack on '
                f'goal "{g}". This is a confirmed vulnerability in the protocol.')
    if llm_v == 'ATTACK' and ofmc_v == 'NO_ATTACK':
        snippet = (f' LLM reasoning: "{r[:160]}..."' if len(r) > 160
                   else (f' LLM reasoning: "{r}"' if r else ''))
        return (f'The LLM flagged goal "{g}" as vulnerable (ATTACK) but OFMC '
                f'found no attack. This is a False Positive — the LLM may have '
                f'over-estimated a threat or misread the protocol logic.{snippet}')
    if llm_v == 'NO_ATTACK' and ofmc_v == 'ATTACK':
        return (f'OFMC detected an attack on goal "{g}" that the LLM missed '
                f'(LLM returned NO_ATTACK). This is a False Negative — a real '
                f'vulnerability was overlooked by the LLM analysis.')
    return 'Comparison could not be completed due to missing data.'
 
 
def time_comparison_note(llm_s, ofmc_ms):
    if pd.isna(llm_s) or pd.isna(ofmc_ms):
        return 'Timing data unavailable.'
    ofmc_s = ofmc_ms / 1000
    ratio  = llm_s / ofmc_s if ofmc_s > 0 else float('inf')
    if ratio >= 2:
        return (f'LLM took {llm_s:.1f}s vs OFMC {ofmc_s:.3f}s '
                f'— LLM was {ratio:.0f}x slower.')
    if ratio < 1:
        return (f'LLM took {llm_s:.1f}s vs OFMC {ofmc_s:.3f}s '
                f'— LLM was faster for this entry.')
    return (f'LLM took {llm_s:.1f}s vs OFMC {ofmc_s:.3f}s '
            f'— comparable execution times.')
 
 
def auto_col_width(ws, min_w=10, max_w=70):
    for col in ws.columns:
        lengths = [len(str(c.value)) if c.value else 0 for c in col]
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            max(min_w, min(max(lengths) + 2, max_w)))
 
 
def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill      = C_HEADER
        cell.font      = F_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = THIN
    ws.row_dimensions[row].height = 30
 
 
def style_data_rows(ws, start=2, zebra=True):
    for r_idx, row in enumerate(ws.iter_rows(min_row=start), start=start):
        for cell in row:
            if zebra and r_idx % 2 == 0:
                cell.fill = C_ALT
            cell.border    = THIN
            cell.alignment = Alignment(wrap_text=True, vertical='top')
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 4.  Load & clean
# ─────────────────────────────────────────────────────────────────────────────
print("Loading data...")
df_llm = pd.read_csv(SUMMARY_CSV)
df_m1  = pd.read_csv(MANUAL_S1)
df_m2  = pd.read_csv(MANUAL_S2)
 
df_llm['protocol_base'] = df_llm['protocol_file'].apply(normalise_llm_protocol)
 
for df_m in [df_m1, df_m2]:
    parsed = df_m['protocol'].apply(parse_manual_protocol)
    df_m['protocol_base'] = parsed.apply(lambda x: x[0])
    df_m['goal_index']    = parsed.apply(lambda x: x[1])
 
df_m1 = df_m1[df_m1['attack'] != 'unknown'].copy()
df_m2 = df_m2[df_m2['attack'] != 'unknown'].copy()
 
print(f"  LLM rows         : {len(df_llm)}")
print(f"  Manual sessions=1: {len(df_m1)}")
print(f"  Manual sessions=2: {len(df_m2)}")
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 5.  Merge & enrich
# ─────────────────────────────────────────────────────────────────────────────
 
def build_comparison(df_llm_in, df_manual_in, session_label):
    left = df_llm_in[[
        'protocol_base', 'llm_goal_index', 'protocol_file',
        'llm_attack', 'llm_goal', 'llm_reasoning',
        'duration_seconds', 'model', 'tokens_used'
    ]].rename(columns={'llm_goal_index': 'goal_index'})
 
    right = df_manual_in[[
        'protocol_base', 'goal_index',
        'attack', 'goal', 'execution-time', 'number-goals'
    ]].rename(columns={
        'attack':         'ofmc_attack',
        'goal':           'ofmc_goal',
        'execution-time': 'ofmc_exec_ms',
        'number-goals':   'ofmc_num_goals',
    })
 
    merged = pd.merge(left, right, on=['protocol_base', 'goal_index'],
                      how='outer', indicator=True)
 
    merged['attack_match'] = merged.apply(
        lambda r: (r['llm_attack'] == r['ofmc_attack'])
                  if pd.notna(r['llm_attack']) and pd.notna(r['ofmc_attack'])
                  else np.nan, axis=1)
 
    merged['ofmc_exec_sec']       = merged['ofmc_exec_ms'] / 1000
    merged['disagreement_type']   = merged.apply(
        lambda r: classify_disagreement(r['llm_attack'], r['ofmc_attack']), axis=1)
    merged['verdict_explanation'] = merged.apply(
        lambda r: verdict_interpretation(
            r['llm_attack'], r['ofmc_attack'],
            r['llm_goal'], r['ofmc_goal'], r['llm_reasoning']), axis=1)
    merged['time_note'] = merged.apply(
        lambda r: time_comparison_note(r['duration_seconds'], r['ofmc_exec_ms']), axis=1)
    merged['session_label'] = session_label
    merged['in_source'] = merged['_merge'].map(
        {'both': 'Both', 'left_only': 'LLM only', 'right_only': 'OFMC only'})
    return merged.drop(columns=['_merge'])
 
 
comp_s1 = build_comparison(df_llm, df_m1, 'Sessions=1')
comp_s2 = build_comparison(df_llm, df_m2, 'Sessions=2')
 
print(f"  Matched (sessions=1): {len(comp_s1[comp_s1['in_source']=='Both'])}")
print(f"  Matched (sessions=2): {len(comp_s2[comp_s2['in_source']=='Both'])}")
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 6.  Statistics
# ─────────────────────────────────────────────────────────────────────────────
 
def compute_stats(comp, label):
    both  = comp[comp['in_source'] == 'Both'].copy()
    total = len(both)
 
    tp = int(((both['llm_attack'] == 'ATTACK')    & (both['ofmc_attack'] == 'ATTACK')).sum())
    tn = int(((both['llm_attack'] == 'NO_ATTACK') & (both['ofmc_attack'] == 'NO_ATTACK')).sum())
    fp = int(((both['llm_attack'] == 'ATTACK')    & (both['ofmc_attack'] == 'NO_ATTACK')).sum())
    fn = int(((both['llm_attack'] == 'NO_ATTACK') & (both['ofmc_attack'] == 'ATTACK')).sum())
 
    matched   = tp + tn
    accuracy  = matched / total * 100 if total else 0
    precision = tp / (tp + fp)  * 100 if (tp + fp) else 0
    recall    = tp / (tp + fn)  * 100 if (tp + fn) else 0
    f1        = (2 * precision * recall / (precision + recall)
                 if (precision + recall) else 0)
 
    avg_llm  = both['duration_seconds'].mean()
    avg_ofmc = both['ofmc_exec_sec'].mean()
 
    return {
        'Session Config':              label,
        'Total Goals Compared':        total,
        'Verdict Match (Agree)':       matched,
        'Verdict Mismatch (Disagree)': total - matched,
        'Accuracy (%)':                round(accuracy,  2),
        'True Positives (TP)':         tp,
        'True Negatives (TN)':         tn,
        'False Positives (FP)':        fp,
        'False Negatives (FN)':        fn,
        'Precision (%)':               round(precision, 2),
        'Recall (%)':                  round(recall,    2),
        'F1 Score (%)':                round(f1,        2),
        'Avg LLM Analysis Time (s)':   round(avg_llm,  2),
        'Avg OFMC Exec Time (s)':      round(avg_ofmc, 3),
        'What TP means':
            'Both LLM and OFMC identified an attack on the same goal — a confirmed vulnerability.',
        'What TN means':
            'Both LLM and OFMC agreed there is no attack — goal is correctly protected.',
        'What FP means':
            'LLM said ATTACK but OFMC found no attack. LLM over-flagged a safe goal (false alarm).',
        'What FN means':
            'OFMC found an ATTACK that LLM missed. A real vulnerability was not caught by the LLM.',
        'Accuracy interpretation':
            ('Strong agreement — LLM and OFMC aligned on most goals.'
             if accuracy >= 75 else
             'Moderate agreement — notable gaps between LLM and OFMC verdicts.'
             if accuracy >= 50 else
             'Weak agreement — LLM verdicts diverge significantly from OFMC.'),
        'Precision interpretation':
            ('When LLM flags an attack it is usually correct.'
             if precision >= 70 else
             'LLM frequently raises attacks that OFMC does not confirm.'
             if precision < 50 else
             'LLM attack flags are correct about half the time.'),
        'Recall interpretation':
            ('LLM captures most of the attacks OFMC finds.'
             if recall >= 70 else
             'LLM misses a meaningful number of real attacks.'
             if recall < 50 else
             'LLM captures roughly half of the attacks OFMC detects.'),
    }
 
 
stats_s1 = compute_stats(comp_s1, 'Sessions = 1')
stats_s2 = compute_stats(comp_s2, 'Sessions = 2')
 
METRIC_COLS = [
    'Session Config', 'Total Goals Compared',
    'Verdict Match (Agree)', 'Verdict Mismatch (Disagree)', 'Accuracy (%)',
    'True Positives (TP)', 'True Negatives (TN)',
    'False Positives (FP)', 'False Negatives (FN)',
    'Precision (%)', 'Recall (%)', 'F1 Score (%)',
    'Avg LLM Analysis Time (s)', 'Avg OFMC Exec Time (s)',
]
INTERP_COLS = [
    'Session Config',
    'What TP means', 'What TN means', 'What FP means', 'What FN means',
    'Accuracy interpretation', 'Precision interpretation', 'Recall interpretation',
]
 
df_stats_num  = pd.DataFrame([{k: v for k, v in s.items() if k in METRIC_COLS}
                               for s in [stats_s1, stats_s2]])
df_stats_text = pd.DataFrame([{k: v for k, v in s.items() if k in INTERP_COLS}
                               for s in [stats_s1, stats_s2]])
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 7.  Per-protocol summary
# ─────────────────────────────────────────────────────────────────────────────
 
def per_protocol_summary(comp):
    both = comp[comp['in_source'] == 'Both'].copy()
    grp  = both.groupby('protocol_base').agg(
        Total_Goals   = ('attack_match', 'count'),
        Goals_Matched = ('attack_match', 'sum'),
        LLM_Attacks   = ('llm_attack',  lambda x: (x == 'ATTACK').sum()),
        OFMC_Attacks  = ('ofmc_attack', lambda x: (x == 'ATTACK').sum()),
        TP            = ('disagreement_type', lambda x: (x.str.startswith('True Positive')).sum()),
        TN            = ('disagreement_type', lambda x: (x.str.startswith('True Negative')).sum()),
        FP            = ('disagreement_type', lambda x: (x.str.startswith('False Positive')).sum()),
        FN            = ('disagreement_type', lambda x: (x.str.startswith('False Negative')).sum()),
        Avg_LLM_s     = ('duration_seconds', 'mean'),
        Avg_OFMC_s    = ('ofmc_exec_sec',    'mean'),
    ).reset_index()
 
    grp['Accuracy (%)'] = (grp['Goals_Matched'] / grp['Total_Goals'] * 100).round(2)
 
    def level(row):
        if row['Accuracy (%)'] == 100:  return 'Full Agreement'
        if row['FP'] > 0 and row['FN'] == 0: return 'LLM Over-flagged (FP only)'
        if row['FN'] > 0 and row['FP'] == 0: return 'LLM Missed Attacks (FN only)'
        if row['FP'] > 0 and row['FN'] > 0:  return 'Mixed Disagreement (FP + FN)'
        if row['Goals_Matched'] > 0:          return 'Partial Agreement'
        return 'Full Disagreement'
 
    def narrative(row):
        parts = []
        if row['TP'] > 0:
            parts.append(f"{int(row['TP'])} goal(s) had attacks confirmed by both tools.")
        if row['TN'] > 0:
            parts.append(f"{int(row['TN'])} goal(s) were agreed safe by both tools.")
        if row['FP'] > 0:
            parts.append(
                f"{int(row['FP'])} goal(s) flagged as ATTACK by LLM but OFMC found none "
                f"— possible LLM over-detection.")
        if row['FN'] > 0:
            parts.append(
                f"{int(row['FN'])} goal(s) had real attacks found by OFMC that LLM missed "
                f"— vulnerabilities not caught by LLM.")
        t_llm  = row['Avg_LLM_s']
        t_ofmc = row['Avg_OFMC_s']
        if pd.notna(t_llm) and pd.notna(t_ofmc) and t_ofmc > 0:
            parts.append(
                f"Avg time: LLM {t_llm:.1f}s vs OFMC {t_ofmc:.3f}s "
                f"({t_llm/t_ofmc:.0f}x difference).")
        return ' '.join(parts)
 
    grp['Agreement Level']  = grp.apply(level, axis=1)
    grp['Protocol Summary'] = grp.apply(narrative, axis=1)
    grp['Avg_LLM_s']  = grp['Avg_LLM_s'].round(2)
    grp['Avg_OFMC_s'] = grp['Avg_OFMC_s'].round(3)
 
    return grp.sort_values('Accuracy (%)').rename(columns={
        'protocol_base': 'Protocol',
        'Total_Goals':   'Total Goals',
        'Goals_Matched': 'Agreed',
        'LLM_Attacks':   'LLM: Attacks Found',
        'OFMC_Attacks':  'OFMC: Attacks Found',
        'Avg_LLM_s':     'Avg LLM Time (s)',
        'Avg_OFMC_s':    'Avg OFMC Time (s)',
    })
 
 
pps1 = per_protocol_summary(comp_s1)
pps2 = per_protocol_summary(comp_s2)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 8.  Goal-level detail
# ─────────────────────────────────────────────────────────────────────────────
 
def goal_detail(comp):
    both = comp[comp['in_source'] == 'Both'].copy()
    both = both.sort_values(['protocol_base', 'goal_index'])
    df = both[[
        'protocol_base', 'goal_index',
        'llm_attack', 'ofmc_attack', 'attack_match',
        'disagreement_type', 'verdict_explanation',
        'llm_goal', 'ofmc_goal',
        'duration_seconds', 'ofmc_exec_ms', 'time_note',
        'llm_reasoning',
    ]].rename(columns={
        'protocol_base':       'Protocol',
        'goal_index':          'Goal Index',
        'llm_attack':          'LLM Verdict',
        'ofmc_attack':         'OFMC Verdict',
        'attack_match':        'Match?',
        'disagreement_type':   'Outcome Category',
        'verdict_explanation': 'What Happened (Explanation)',
        'llm_goal':            'LLM Goal Description',
        'ofmc_goal':           'OFMC Goal Description',
        'duration_seconds':    'LLM Duration (s)',
        'ofmc_exec_ms':        'OFMC Exec Time (ms)',
        'time_note':           'Timing Comparison',
        'llm_reasoning':       'LLM Reasoning',
    })
    df['Match?'] = df['Match?'].map({True: 'YES', False: 'NO', np.nan: '-'})
    return df
 
 
gd1 = goal_detail(comp_s1)
gd2 = goal_detail(comp_s2)
disc1 = gd1[gd1['Match?'] == 'NO'].copy()
disc2 = gd2[gd2['Match?'] == 'NO'].copy()
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 9.  Confusion matrix
# ─────────────────────────────────────────────────────────────────────────────
 
def confusion_df(s):
    return pd.DataFrame({
        'OFMC Verdict \\ LLM Verdict': ['OFMC: ATTACK', 'OFMC: NO_ATTACK'],
        'LLM: ATTACK':    [s['True Positives (TP)'],  s['False Positives (FP)']],
        'LLM: NO_ATTACK': [s['False Negatives (FN)'], s['True Negatives (TN)']],
    })
 
conf1 = confusion_df(stats_s1)
conf2 = confusion_df(stats_s2)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 10.  Write to Excel
# ─────────────────────────────────────────────────────────────────────────────
print("\nWriting Excel report...")
 
with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as writer:
    pd.DataFrame().to_excel(writer, sheet_name='README', index=False)
    df_stats_num.to_excel(writer,  sheet_name='Summary - Metrics',        index=False)
    df_stats_text.to_excel(writer, sheet_name='Summary - Interpretation', index=False)
    conf1.to_excel(writer,  sheet_name='Confusion Matrix S1', index=False)
    conf2.to_excel(writer,  sheet_name='Confusion Matrix S2', index=False)
    pps1.to_excel(writer,   sheet_name='Per-Protocol S1',     index=False)
    pps2.to_excel(writer,   sheet_name='Per-Protocol S2',     index=False)
    disc1.to_excel(writer,  sheet_name='Discrepancies S1',    index=False)
    disc2.to_excel(writer,  sheet_name='Discrepancies S2',    index=False)
    gd1.to_excel(writer,    sheet_name='All Goals S1',        index=False)
    gd2.to_excel(writer,    sheet_name='All Goals S2',        index=False)
 
 
# ─────────────────────────────────────────────────────────────────────────────
# 11.  Style workbook
# ─────────────────────────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT_PATH)
 
# ── README ────────────────────────────────────────────────────────────────────
ws = wb['README']
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 72
 
T16  = Font(bold=True, size=16, color='1F4E79')
T12  = Font(bold=True, size=12, color='2E75B6')
T11  = Font(bold=True, size=11)
BODY = Font(size=11)
 
def rw(ws, row, col, text, font, col2=None, text2=None, font2=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font; c.alignment = Alignment(wrap_text=True, vertical='top')
    if col2 and text2:
        c2 = ws.cell(row=row, column=col2, value=text2)
        c2.font = font2 or BODY
        c2.alignment = Alignment(wrap_text=True, vertical='top')
 
rw(ws, 2,  2, 'Protocol Security Comparison: LLM vs OFMC', T16)
rw(ws, 4,  2, 'Purpose', T12)
rw(ws, 5,  2,
   'This workbook compares security verdicts from a Large Language Model (LLM, DeepSeek) '
   'against manually-run OFMC model-checking results across 130 cryptographic protocol families. '
   'For each protocol security goal it shows whether both tools agreed, where they differed, '
   'why, and how long each approach took.', BODY)
 
rw(ws, 7,  2, 'Column Glossary', T12)
glossary = [
    ('LLM Verdict',              'The verdict produced by the LLM: ATTACK or NO_ATTACK.'),
    ('OFMC Verdict',             'The ground-truth verdict from the OFMC model-checking tool.'),
    ('Match?',                   'YES = both tools gave the same verdict. NO = they disagreed.'),
    ('Outcome Category',         'TP / TN / FP / FN — see legend below.'),
    ('What Happened (Explanation)',
     'Plain-English description of the agreement or disagreement for that specific goal.'),
    ('Timing Comparison',        'Sentence comparing LLM analysis time vs OFMC execution time.'),
    ('LLM Reasoning',            "The LLM's own justification for its verdict."),
    ('Protocol Summary',         'Per-protocol narrative: what matched, what differed, timing gap.'),
    ('Agreement Level',          'High-level label: Full Agreement / Over-flagged / Missed Attacks / Mixed.'),
]
for i, (col_name, desc) in enumerate(glossary, start=8):
    rw(ws, i, 2, col_name, T11, 3, desc, BODY)
 
rw(ws, 18, 2, 'Outcome Category Legend', T12)
legend_items = [
    (19, C_GREEN,      'True Positive (TP)',
     'Both LLM and OFMC detected an attack. Real vulnerability confirmed by both tools.'),
    (20, C_BLUE_LIGHT, 'True Negative (TN)',
     'Both agreed there is no attack. The protocol goal is correctly protected.'),
    (21, C_ORANGE,     'False Positive (FP)',
     'LLM said ATTACK; OFMC found nothing. LLM over-detected — possible false alarm.'),
    (22, C_RED,        'False Negative (FN)',
     'OFMC found an ATTACK; LLM said NO_ATTACK. LLM missed a real vulnerability.'),
]
for row, fill, label, desc in legend_items:
    c = ws.cell(row=row, column=2, value=label)
    c.fill = fill; c.font = T11; c.border = THIN
    c.alignment = Alignment(vertical='center')
    d = ws.cell(row=row, column=3, value=desc)
    d.font = BODY; d.alignment = Alignment(wrap_text=True, vertical='top')
 
rw(ws, 24, 2, 'Sheet Guide', T12)
guide = [
    ('README',                   'This sheet — glossary and legend.'),
    ('Summary - Metrics',        'Accuracy, precision, recall, F1, timing — numbers only.'),
    ('Summary - Interpretation', 'Plain-English meaning of every metric for both session configs.'),
    ('Confusion Matrix S1/S2',   'TP/TN/FP/FN in a 2x2 grid, colour-coded per cell type.'),
    ('Per-Protocol S1/S2',       'One row per protocol: counts, agreement level, and a narrative summary of differences.'),
    ('Discrepancies S1/S2',      'ONLY rows where LLM and OFMC disagreed — most useful for pinpointing errors.'),
    ('All Goals S1/S2',          'Every individual goal: side-by-side verdicts, explanations, and timing.'),
]
for i, (sh, desc) in enumerate(guide, start=25):
    rw(ws, i, 2, sh, T11, 3, desc, BODY)
 
 
# ── Summary - Metrics ─────────────────────────────────────────────────────────
ws = wb['Summary - Metrics']
style_header_row(ws)
style_data_rows(ws)
ws.freeze_panes = 'B2'
auto_col_width(ws)
headers = [c.value for c in ws[1]]
for col_name in ['Accuracy (%)', 'Precision (%)', 'Recall (%)', 'F1 Score (%)']:
    if col_name not in headers: continue
    ci = headers.index(col_name) + 1
    for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ci, max_col=ci):
        for c in cell:
            try:
                v = float(c.value)
                c.fill = C_GREEN if v >= 75 else (C_YELLOW if v >= 50 else C_RED)
                c.font = F_BOLD
            except (TypeError, ValueError):
                pass
 
 
# ── Summary - Interpretation ──────────────────────────────────────────────────
ws = wb['Summary - Interpretation']
style_header_row(ws)
for col in ws.columns:
    ws.column_dimensions[get_column_letter(col[0].column)].width = 65
for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    for cell in row:
        cell.border = THIN
        cell.font   = F_NORMAL
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r_idx].height = 60
 
 
# ── Confusion matrices ────────────────────────────────────────────────────────
for sh in ['Confusion Matrix S1', 'Confusion Matrix S2']:
    ws = wb[sh]
    style_header_row(ws)
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        for c_idx, cell in enumerate(row, start=1):
            cell.border    = THIN
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c_idx == 1:
                cell.font = F_BOLD; continue
            try: int(cell.value)
            except (TypeError, ValueError): continue
            if r_idx == 1 and c_idx == 2:   cell.fill = C_GREEN        # TP
            elif r_idx == 2 and c_idx == 3: cell.fill = C_BLUE_LIGHT   # TN
            elif r_idx == 1 and c_idx == 3: cell.fill = C_RED          # FN
            elif r_idx == 2 and c_idx == 2: cell.fill = C_ORANGE       # FP
            cell.font = Font(bold=True, size=14)
    auto_col_width(ws)
 
 
# ── Per-Protocol sheets ───────────────────────────────────────────────────────
for sh in ['Per-Protocol S1', 'Per-Protocol S2']:
    ws = wb[sh]
    style_header_row(ws)
    ws.freeze_panes = 'B2'
    headers = [c.value for c in ws[1]]
 
    def ci(name): return headers.index(name) + 1 if name in headers else None
 
    acc_c   = ci('Accuracy (%)')
    agree_c = ci('Agreement Level')
    narr_c  = ci('Protocol Summary')
 
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border    = THIN
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if acc_c and cell.column == acc_c:
                try:
                    v = float(cell.value)
                    cell.fill = C_GREEN if v == 100 else (C_YELLOW if v >= 50 else C_RED)
                    cell.font = F_BOLD
                except (TypeError, ValueError): pass
            elif agree_c and cell.column == agree_c:
                v = str(cell.value)
                if   'Full Agreement'  in v: cell.fill = C_GREEN;  cell.font = F_BOLD
                elif 'Over-flagged'    in v: cell.fill = C_ORANGE; cell.font = F_BOLD
                elif 'Missed'          in v: cell.fill = C_RED;    cell.font = F_BOLD
                elif 'Mixed'           in v: cell.fill = C_PURPLE; cell.font = F_BOLD
                elif 'Partial'         in v: cell.fill = C_YELLOW; cell.font = F_BOLD
                else:                        cell.fill = C_GREY
            elif narr_c and cell.column == narr_c:
                cell.font = F_ITALIC
            elif r_idx % 2 == 0:
                cell.fill = C_ALT
 
    if narr_c:
        ws.column_dimensions[get_column_letter(narr_c)].width = 75
    auto_col_width(ws)
 
 
# ── All Goals & Discrepancy sheets ────────────────────────────────────────────
for sh in ['All Goals S1', 'All Goals S2', 'Discrepancies S1', 'Discrepancies S2']:
    ws = wb[sh]
    style_header_row(ws)
    ws.freeze_panes = 'A2'
    headers = [c.value for c in ws[1]]
 
    def ci(name): return headers.index(name) + 1 if name in headers else None
 
    match_c  = ci('Match?')
    llm_v_c  = ci('LLM Verdict')
    ofmc_v_c = ci('OFMC Verdict')
    cat_c    = ci('Outcome Category')
    exp_c    = ci('What Happened (Explanation)')
    time_c   = ci('Timing Comparison')
 
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border    = THIN
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font      = F_NORMAL
 
            if match_c and cell.column == match_c:
                if str(cell.value) == 'YES':
                    cell.fill = C_GREEN; cell.font = Font(bold=True, color='375623')
                elif str(cell.value) == 'NO':
                    cell.fill = C_RED;   cell.font = Font(bold=True, color='9C0006')
            elif llm_v_c and cell.column == llm_v_c and str(cell.value) == 'ATTACK':
                cell.fill = C_ORANGE; cell.font = F_BOLD
            elif ofmc_v_c and cell.column == ofmc_v_c and str(cell.value) == 'ATTACK':
                cell.fill = C_ORANGE; cell.font = F_BOLD
            elif cat_c and cell.column == cat_c:
                v = str(cell.value)
                if   'True Positive'  in v: cell.fill = C_GREEN;      cell.font = Font(bold=True, color='375623')
                elif 'True Negative'  in v: cell.fill = C_BLUE_LIGHT; cell.font = F_BOLD
                elif 'False Positive' in v: cell.fill = C_ORANGE;     cell.font = Font(bold=True, color='833C00')
                elif 'False Negative' in v: cell.fill = C_RED;        cell.font = Font(bold=True, color='9C0006')
            elif exp_c  and cell.column == exp_c:  cell.font = F_ITALIC
            elif time_c and cell.column == time_c: cell.font = F_ITALIC
            elif r_idx % 2 == 0:
                cell.fill = C_ALT
 
    for name, width in [('What Happened (Explanation)', 65),
                         ('LLM Reasoning', 65),
                         ('LLM Goal Description', 42),
                         ('OFMC Goal Description', 42),
                         ('Timing Comparison', 48)]:
        if name in headers:
            ws.column_dimensions[get_column_letter(headers.index(name)+1)].width = width
    auto_col_width(ws)
 
wb.save(OUTPUT_PATH)
print(f"Report saved to: {OUTPUT_PATH}")
 
# ─────────────────────────────────────────────────────────────────────────────
# 12.  Console summary
# ─────────────────────────────────────────────────────────────────────────────
SEP = "=" * 65
print(f"\n{SEP}\nQUICK RESULTS SUMMARY\n{SEP}")
for s in [stats_s1, stats_s2]:
    print(f"\n{s['Session Config']}")
    print(f"  Goals compared    : {s['Total Goals Compared']}")
    print(f"  Agreed            : {s['Verdict Match (Agree)']}  |  "
          f"Disagreed: {s['Verdict Mismatch (Disagree)']}")
    print(f"  Accuracy          : {s['Accuracy (%)']:.2f}% — {s['Accuracy interpretation']}")
    print(f"  TP={s['True Positives (TP)']}  TN={s['True Negatives (TN)']}  "
          f"FP={s['False Positives (FP)']}  FN={s['False Negatives (FN)']}")
    print(f"  Precision         : {s['Precision (%)']:.2f}% — {s['Precision interpretation']}")
    print(f"  Recall            : {s['Recall (%)']:.2f}% — {s['Recall interpretation']}")
    print(f"  F1 Score          : {s['F1 Score (%)']:.2f}%")
    print(f"  Avg LLM time      : {s['Avg LLM Analysis Time (s)']}s")
    print(f"  Avg OFMC time     : {s['Avg OFMC Exec Time (s)']}s")
    print(f"  FP note: {s['What FP means']}")
    print(f"  FN note: {s['What FN means']}")
 
print(f"\nDiscrepancy rows: S1={len(disc1)}, S2={len(disc2)}")
print(f"Total Excel sheets: {len(wb.sheetnames)}")
 